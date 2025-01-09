# -*- coding: utf-8 -*-
"""
-- ============================================================================
-- Descripción....: 
	(1) Generación de Reporte Excel: 
        Formato 2.2 - Mediciones mensuales transformadores AT-MT       
    
-- Elabora........: ArBR (arcebrito@gmail.com)
-- Fecha..........: 2019-02-05
-- ============================================================================
"""

import time
import utlpy
import shutil
import pymysql
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from threading import Thread

ZERO_ROW            = 8
GENERAR_VACIOS      = True
REPOXLS_PATH        = ""
PR_NAME             = "F22MMTATMT"
LOGO_CFE_DIST_PATH  = "TEMPLATE\logo_cfe_distribucion.png"
TEMPLATE_PATH       = "TEMPLATE\F22MMTatmt_template.xlsx"
REPOXLS_BASE_PATH   =  'E:\\repoxls' if utlpy.drive_exists('e') else 'c:\\repoxls'

CN = {"host": "10.4.22.84", "user":  "ClusterBR", "passwd": "XYZ123...", "database": "apcc"}

def fn_get_ds_bancos(connection, cveDivision, cveZona) :
    qry =""" 
        select distinct b.claveSubestacion, b.numerobanco, s.Nombre, bc.bcapacidad
            from apcc.tbcatbanco b
            inner join apcc.tbcatsubestacion s on
            s.ClaveDivision = b.claveDivision and s.ClaveZona = b.claveZona and s.Clave = b.claveSubestacion
            inner join apcc.catalogo_equipos e on 
            e.clave_division = s.ClaveDivision and e.cveZona = s.ClaveZona and e.clave_subestacion = s.Clave 
               and e.numero_de_transformador = b.numerobanco
            inner join bancos bc on
            	e.clave_division = bc.bdiv and e.cveZona = bc.bzona and e.clave_subestacion = bc.bsub 
            	 and e.numero_de_transformador = concat('0', bc.bbanco)
            where b.claveDivision = %s and b.claveZona = %s and e.tipo_equipo = 'TRANSFORMADOR'
            group by b.nodeserie
            order by claveSubestacion, numerobanco;
          """
    
    params = (cveDivision, cveZona)
    return utlpy.mysql_fecthall_dict(connection, qry, params)

def fn_get_ds_mediciones(connection, cveDivision, cveZona, cveSub, banco, anio) :
    qry =""" 
        select descripcion, 
        	max(case when bhmes = 'ENE' then magnitud else 0 end) ENE,
        	max(case when bhmes = 'FEB' then magnitud else 0 end) FEB,
            max(case when bhmes = 'MAR' then magnitud else 0 end) MAR,
            max(case when bhmes = 'ABR' then magnitud else 0 end) ABR,
            max(case when bhmes = 'MAY' then magnitud else 0 end) MAY,
            max(case when bhmes = 'JUN' then magnitud else 0 end) JUN,
            max(case when bhmes = 'JUL' then magnitud else 0 end) JUL,
            max(case when bhmes = 'AGO' then magnitud else 0 end) AGO,
            max(case when bhmes = 'SEP' then magnitud else 0 end) SEP,
            max(case when bhmes = 'OCT' then magnitud else 0 end) OCT,
            max(case when bhmes = 'NOV' then magnitud else 0 end) NOV,
            max(case when bhmes = 'DIC' then magnitud else 0 end) DIC    
         from
         (
          select bhmes, bhdem as magnitud, 'Demanda (kW)' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
          union all select bhmes, bhener as magnitud, 'Energía (kWh)' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
          union all select bhmes, bhreac as magnitud, 'Reactivos (kVARh)' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
          union all select bhmes, bhfp as magnitud, 'Factor de Potencia' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
          union all select bhmes, bhdemreac as magnitud, 'Demanda Reactiva (kVAR)' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
          union all select bhmes, bhdemmed as magnitud, 'Demanda Media' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
          union all select bhmes, bhfaccar as magnitud, 'Factor de Carga' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
          union all select bhmes, bhfu as magnitud, 'Factor de Utilización' as descripcion 
              from apcc.dm_bancohist where bhdiv='[DIV]' and bhzona='[ZN]' and bhsub='[SUB]' and bhban='[NB]' and bhultact='[ANIO]'
         ) src
         group by descripcion
          """
    qry = qry.replace("[DIV]", cveDivision)
    qry = qry.replace("[ZN]", cveZona)
    qry = qry.replace("[SUB]", cveSub)    
    qry = qry.replace("[NB]", str(utlpy.to_number(banco)))
    qry = qry.replace("[ANIO]", anio)
    
    return utlpy.mysql_fecthall(connection, qry)

def insert_logo_cfe(ws) :    
    img = openpyxl.drawing.image.Image(LOGO_CFE_DIST_PATH)
    ws.cell(row=2, column=2)
    ws.add_image(img)
    return

def style_cell_medicion(cell) :
    utlpy.pyxl_center_borderall_fill_cell(cell, fill=False, center=False, border=False, fgColor="FFFFFF")    
    return

def style_cell_dark(cell) :
    ft = Font(bold=True)
    cell.font = ft;
    utlpy.pyxl_center_borderall_fill_cell(cell, fill=True, center=False, border=False, fgColor="C0C0C0")  
    return

def fn_create_banco_first_row(ws, rs_zona, rs_banco, idx_banco) :    
    row = ZERO_ROW + ( (idx_banco-1) * 10)
    
    numerobanco = str(utlpy.to_number(rs_banco["numerobanco"]))
    banco_info = "Banco {} Cap: {} MVA".format(numerobanco, rs_banco["bcapacidad"])
    style_cell_dark(ws.cell(row = row, column = 1, value = rs_zona["Abreviatura"]))
    style_cell_dark(ws.cell(row = row, column = 2, value = rs_banco["claveSubestacion"]))
    style_cell_dark(ws.cell(row = row, column = 3, value = banco_info))
    
    for x in range(4, 16) :
        style_cell_dark(ws.cell(row = row, column = x, value = ""))
    #end-if
    return

def fn_create_banco_measurement_rows(ws, rs_zona, rs_banco, ds_mediciones, idx_banco) :    
    row = ZERO_ROW + ( (idx_banco-1) * 10)  + 1   
    
    for x in range(0, 8) :
        style_cell_medicion(ws.cell(row = row + x, column = 2, value = ds_mediciones[x][0]))
        style_cell_medicion(ws.cell(row = row + x, column = 3, value = ""))        

        style_cell_medicion(ws.cell(row = row + x, column = 4, value = float(ds_mediciones[x][1])))        
        style_cell_medicion(ws.cell(row = row + x, column = 5, value = ds_mediciones[x][2]))
        style_cell_medicion(ws.cell(row = row + x, column = 6, value = ds_mediciones[x][3]))
        style_cell_medicion(ws.cell(row = row + x, column = 7, value = ds_mediciones[x][4]))
        style_cell_medicion(ws.cell(row = row + x, column = 8, value = ds_mediciones[x][5]))
        style_cell_medicion(ws.cell(row = row + x, column = 9, value = ds_mediciones[x][6]))
        style_cell_medicion(ws.cell(row = row + x, column = 10, value = ds_mediciones[x][7]))
        style_cell_medicion(ws.cell(row = row + x, column = 11, value = ds_mediciones[x][8]))
        style_cell_medicion(ws.cell(row = row + x, column = 12, value = ds_mediciones[x][9]))
        style_cell_medicion(ws.cell(row = row + x, column = 13, value = ds_mediciones[x][10]))
        style_cell_medicion(ws.cell(row = row + x, column = 14, value = ds_mediciones[x][11]))
        style_cell_medicion(ws.cell(row = row + x, column = 15, value = ds_mediciones[x][12]))
    #end-if
    return


def fn_create_rpt_by_zona(connection, anio, cveDivision, wb, rs_zona) :       
    
    cveZona = rs_zona["claveZona"]        
    ws = wb.copy_worksheet(wb["ZONA"])
    ws.title = "{}{}".format(cveDivision, cveZona)
    
    insert_logo_cfe(ws)
    nombre_division = utlpy.fn_get_division_only(rs_zona["NombreDivision"])
    cell_div = ws.cell(row = 2, column = 1)
    cell_zon = ws.cell(row = 3, column = 1)
    cell_anio = ws.cell(row = 5, column = 1)    
    cell_div.value = cell_div.value.replace("[NOMBRE_DIVISION]", nombre_division)
    cell_zon.value = cell_zon.value.replace("[NOMBRE_ZONA]", rs_zona["Titulo"])
    cell_anio.value = cell_anio.value.replace("[ANIO]", anio)
    
    idx_banco = 0
    ds_bancos = fn_get_ds_bancos(connection, cveDivision, cveZona)
    for rs in ds_bancos :        
        utlpy.println("{}-{} - Banco:{}".format(cveDivision, cveZona, rs["Nombre"]))        
        ds_mediciones = fn_get_ds_mediciones(connection, cveDivision, cveZona, rs["claveSubestacion"], rs["numerobanco"], anio)

        if GENERAR_VACIOS :
            idx_banco = idx_banco + 1
            fn_create_banco_first_row(ws, rs_zona, rs, idx_banco)
            if ds_mediciones:
                fn_create_banco_measurement_rows(ws, rs_zona, rs, ds_mediciones, idx_banco)
        else :
            if ds_mediciones:
                idx_banco = idx_banco + 1
                fn_create_banco_first_row(ws, rs_zona, rs, idx_banco)            
                fn_create_banco_measurement_rows(ws, rs_zona, rs, ds_mediciones, idx_banco)
    #end-for
    
    return

def proc_genera_xls_by_zona (anio, cveDivision, cveZona, rs_zona) :
    
    start_time = time.time()    
    prid = utlpy.btc_gen_prid(PR_NAME, "{}{}{}".format(cveDivision, cveZona, anio))
    connection = pymysql.connect(host=CN["host"], user=CN["user"], passwd=CN["passwd"], database=CN["database"]) 
    dest_fname = "{}\{}_{}{}{}.xlsx".format(REPOXLS_PATH, PR_NAME, anio, cveDivision, cveZona)
    
    utlpy.btc_insert(connection, prid, PR_NAME, "INICIADO", "", cveDivision, cveZona, "", "", anio, "*")
    try:
        utlpy.println("proc_genera_xls_by_zona:{} thread started & running (...)".format(cveZona))
        
        shutil.copy(TEMPLATE_PATH, dest_fname)
        wb = load_workbook(dest_fname)
        
        fn_create_rpt_by_zona(connection, anio, cveDivision, wb, rs_zona)
        
        wb.remove(wb["ZONA"])
        wb.save(dest_fname) 
        
        elapsed_time = time.time() - start_time
        elapsed_time_fmt = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))
        
        message = "DURACION {}".format(str(elapsed_time_fmt))
        utlpy.println("proc_genera_xls_by_zona:{} thread completed. Total time: {}".format(cveZona, elapsed_time_fmt))
        utlpy.btc_insert(connection, prid, PR_NAME, "COMPLETADO", message, cveDivision, cveZona, "", "", anio, "*")
        
    except Exception as e:
        message = str(e)
        utlpy.println('proc_genera_xls_by_zona.error>>> ' + message)        
        utlpy.btc_insert(connection, prid, PR_NAME, "ERROR", message, cveDivision, cveZona, "", "", anio, "*")
        
    finally:
        connection.close()

    return

def proc_genera_xls_by_div(cveDivision, anio) :
    
    
    utlpy.println("proc_genera_xls_by_div: >>> main thread started >>>")
    
    connection = None
    try:
        connection = pymysql.connect(host=CN["host"], user=CN["user"], passwd=CN["passwd"], database=CN["database"]) 
        lst_threads = []
        ds_zonas = utlpy.fn_get_ds_zonas(connection, cveDivision)
        connection.close()
        connection = None
        
        for rs_zona in ds_zonas :
            cveZona = rs_zona["claveZona"]
            t = Thread(target = proc_genera_xls_by_zona, args = (anio, cveDivision, cveZona, rs_zona))
            lst_threads.append(t)
        
        [t.start() for t in lst_threads]
        [t.join() for t in lst_threads]
            
        zip_file_name = "{}_{}{}".format(PR_NAME, anio, cveDivision)
        utlpy.create_parent_zip_from_dir(REPOXLS_PATH, zip_file_name)
        
    except Exception as e:
        utlpy.println('proc_genera_xls_by_div.error>>> {}'.format(str(e)))
    finally:
        if connection:
            connection.close()
            
    print("proc_genera_xls_by_div: >>> main thread completed >>>")
    return


##########################
# main
##########################

if __name__ == '__main__' :
    
    anio = "2018"
    lst_divisiones = ['DA','DB','DC','DD','DF','DG','DJ','DK','DL','DM','DN','DP','DU','DV','DW','DX']
    for division in lst_divisiones :
        REPOXLS_PATH = "{}\{}\{}\{}".format(REPOXLS_BASE_PATH, PR_NAME, anio,  division)
        utlpy.create_dir_if_not_exists(REPOXLS_PATH)
        proc_genera_xls_by_div(division, anio)
        