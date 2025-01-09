# -*- coding: utf-8 -*-
"""
-- ============================================================================
-- Descripción....: Helper functions for python
    
-- Elabora........: ArBR (arcebrito@gmail.com)
-- Fecha..........: 2019-01-29
-- ============================================================================
"""

import os
import sys
import math
import time
import socket
import shutil
import datetime
import pymysql
import logging
import platform

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment

def println(str_data):
    print( "{}\n".format(str_data))
    return 


def get_host_name() :
    result = ""
    try:
        result = socket.gethostname()
    except:
        result = ""    
    return result

def to_int(x) :
    result = x
    try:
        result = int(x)
    except:
        result =  x
    return result

def to_int_or_def(x, default) :
    result = default
    try:
        result = int(x)
    except:
        result =  default
    return result	
def to_number(x) :
    result = x
    try:
        result = float(x) if '.' in x else int(x)
    except:
        result =  x
    return result

def truncate(number, digits):
    result = -1.0
    if not number is None:
        try:
            stepper = pow(10.0, digits)
            result = math.trunc(stepper * number) / stepper
        except:
            result = -1.0
    return result				 

def drive_exists(letter) :
    return "Windows" in platform.system() and os.system("vol %s: 2>nul>nul" % (letter)) == 0

def create_dir_if_not_exists(path) :    
    if not os.path.exists(path):
        os.makedirs(path)
    return os.path.exists(path)

def smes(imes):
    return ['','ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC'][imes]

def imes(smes):
    return ['','ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC'].index(smes)

def create_parent_zip_from_dir(source_path, file_name_without_extension) :
    if os.path.exists(source_path) :
        parent_dir = os.path.dirname(source_path)
        destination_file_path_name = "{}\{}".format(parent_dir, file_name_without_extension)
        shutil.make_archive(destination_file_path_name, 'zip', source_path)
    return

def current_time():
    return time.time()

def current_time_fmt():
    return time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
    
def elapsed_time(start_time): 
     return time.time() - start_time

# "%Hrs:%Mts:%Segs"
def elapsed_time_fmt(start_time) :
    end_time = time.time()
    temp = end_time - start_time
    hours = temp // 3600
    temp = temp - 3600 * hours
    minutes = temp // 60
    seconds = temp - 60 * minutes    
    return "{}:{}:{}".format(hours, math.ceil(minutes), math.ceil(seconds))

def mysql_fecthall(connection, qry, params = None):    
    result = None
    cursor = None
    try:
        cursor = pymysql.cursors.Cursor(connection)
        if params:
            cursor.execute(qry, params)
        else:
            cursor.execute(qry)
        result = cursor.fetchall()
    except Exception as e:
        println('mysql_fecthall.error>>> {}'.format(str(e)))
    finally:
        if cursor:
            cursor.close()        
    return result

def mysql_fecthall_dict(connection, qry, params):
    result = None
    cursor = None
    try:        
        cursor = pymysql.cursors.DictCursor(connection)
        cursor.execute(qry, params)       
        result = cursor.fetchall()
    except Exception as e:
        println('mysql_fecthall_dict.error>>> {}'.format(str(e)))
    finally:
        if cursor:
            cursor.close()        
    return result

def mysql_fecthone_dict(connection, qry, params):
    result = None
    cursor = None
    try:
        cursor = pymysql.cursors.DictCursor(connection)
        cursor.execute(qry, params)       
        result = cursor.fetchone()
    except Exception as e:
        println('mysql_fecthone_dict.error>>> {}'.format(str(e)))
    finally:
        if cursor:
            cursor.close()
    return result

def pyxl_center_borderall_fill_cell(cell, fill=False, center=False, border=False, fgColor='FFFFFF', borderColor='000000') :
    
    borderStyle = Border(left=Side(border_style='thin', color=borderColor),
                    right=Side(border_style='thin', color=borderColor),
                    top=Side(border_style='thin', color=borderColor),
                    bottom=Side(border_style='thin', color=borderColor))
    if (border) :
        cell.border = borderStyle
    if (center) :
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if (fill) : 
        cell.fill = PatternFill("solid", fgColor=fgColor)
        
    return


def btc_insert(connection, IDProceso, Proceso, Estatus, Mensaje, Division="", Zona="", Subestacion="", Circuito="", Anio="", Mes=""):
    qry = """ insert into siapcc_btc_procesos(IDProceso, Fecha, Proceso, Estatus, Mensaje,
               	Division, Zona, Subestacion, Circuito, Anio, Mes, IPSolicitud)
               values (%s, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);"""
    
    result = None
    cursor = None
    try:
        IPSolicitud = get_host_name()        
        Fecha = datetime.datetime.now()
        params = (IDProceso, Fecha, Proceso, Estatus, Mensaje, Division, Zona, Subestacion, Circuito, Anio, Mes, IPSolicitud)
        cursor = pymysql.cursors.Cursor(connection)
        cursor.execute(qry, params)
        connection.commit()
        result = cursor.fetchone()
    except Exception as e:
        println('btc_insert.error>>> {}'.format(str(e)))
    finally:
        if cursor:
            cursor.close()
    return result

def btc_gen_prid(proc_name, key):
    now = datetime.datetime.now()
    prid = "{}{}{}".format(now.strftime("%Y%m%d-%H%M%S"), proc_name, key)
    return prid

def fn_get_ds_zonas(connection, cveDivision) :
    qry =  """ select distinct Clave as claveZona, Abreviatura, Nombre, Titulo, NombreDivision
                 from apcc.tbcatzona c
                 inner join apcc.catalogo_equipos e on
                 	e.clave_division = c.ClaveDivision and e.cveZona = c.Clave
                 where ClaveDivision = %s
                 order by Clave asc
            """
    params = (cveDivision)
    return mysql_fecthall_dict(connection, qry, params)


def fn_get_division_only(large_division):
    result = large_division or "*"
    if large_division and "DIVISION" in large_division:
        result = large_division.split("DIVISION", 1)[1]
        result = result.strip()
    return result

def init_log(log, f_name, formatter):
    log_file = logging.FileHandler(f_name, mode="a", encoding="UTF8")
    log_file.setLevel(logging.DEBUG)       
    log_file.setFormatter(formatter)
    
    log_console = logging.StreamHandler(stream=sys.stdout)
    log_console.setLevel(logging.DEBUG)
    log_console.setFormatter(formatter)
    
    log.addHandler(log_file)
    log.addHandler(log_console)    
    return log

def get_log(log_path, log_name) :
    
    log = None
    txt = '[%(asctime)s] %(levelname)8s --- %(message)s (%(filename)s:%(lineno)s)'
    if create_dir_if_not_exists(log_path) :
        formatter = logging.Formatter(txt, datefmt='%Y-%m-%d %H:%M:%S')
        f_name = "{}\{}x.log".format(log_path, log_name)
        log = logging.getLogger(log_name)
        log.setLevel(logging.DEBUG)      
        
        if sys.platform == "linux" or sys.platform == "linux2":            
            if not log.handlers:
                init_log(log, f_name, formatter)
        elif sys.platform == "win32":
            if not log.handlers:
                init_log(log, f_name, formatter)
        #endif platform  
    #end-if-create_dir
    return log
    
