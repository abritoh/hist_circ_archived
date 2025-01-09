# -*- coding: utf-8 -*-
"""
-- ============================================================================
-- Descripción....: 
	(1) Generación de Reporte Excel: 
        Formato 1 - Demandas Horarias Integradas -TR del Servicio Público
    
-- Elabora........: ArBR (arcebrito@gmail.com)
-- Fecha..........: 2019-01-28
-- ============================================================================
"""

import time
import utlpy
import shutil
import datetime
import pymysql
import openpyxl
from openpyxl import load_workbook
from threading import Thread

ZERO_ROW            = 12
ZERO_COL_BANCO      = 5
PR_NAME             = "F1DHITRspf"
LOGO_CENACE_PATH    = "TEMPLATE\logo_cenace.png"
TEMPLATE_PATH       = "TEMPLATE\F1DHITRspf_template.xlsx"
REPOXLS_BASE_PATH   =  'E:\\repoxls' if utlpy.drive_exists('e') else 'c:\\repoxls'

CN = {"host": "10.4.22.84", "user":  "ClusterBR", "passwd": "XYZ123...", "database": "apcc"}

#select bancos en subestacion
def fn_get_ds_bancos(connection, cveDivision, cveZona) :
    qry =(" select distinct s.Nombre, b.nodeserie,"
          "    concat(b.tensionatkvdeoperacion, '/', b.tensionbtkvdeoperacion) as relacionTension,"
          "    b.numerobanco, capacidad3mva,"
          "    bc.bnumcir as numCirc, b.claveSubestacion, e.id_catalogo_equipos"
          " from apcc.tbcatbanco b"
          " inner join apcc.tbcatsubestacion s on"
          " 	s.ClaveDivision = b.claveDivision and s.ClaveZona = b.claveZona and s.Clave = b.claveSubestacion"
          " inner join apcc.catalogo_equipos e on "
          " 	e.clave_division = s.ClaveDivision and e.cveZona = s.ClaveZona and e.clave_subestacion = s.Clave "
          "     and e.numero_de_transformador = b.numerobanco"
          "  inner join bancos bc on"
          "   	e.clave_division = bc.bdiv and e.cveZona = bc.bzona and e.clave_subestacion = bc.bsub "
          "       and e.numero_de_transformador = concat('0', bc.bbanco)"
          " where b.claveDivision = %s and b.claveZona = %s and e.tipo_equipo = 'TRANSFORMADOR'"
          " group by b.nodeserie"
          " order by claveSubestacion, numerobanco;")
    
    params = (cveDivision, cveZona)
    return utlpy.mysql_fecthall_dict(connection, qry, params)

def fn_get_ds_medhr_by_date_range(connection, id_equipo, iso_date1, iso_date2) :
    qry = (" select concat(hr.fecha, ' ', hr.hora) as isofecha, hr.kw, hr.fecha"
           " from apcc.historicos_parametros_electricos_hora hr"
           " inner join apcc.catalogo_equipos e on e.id_catalogo_equipos = hr.id_equipo"
           " where id_catalogo_equipos = %s and fecha between %s and %s"
           " order by isofecha asc;")
    
    params = (id_equipo, iso_date1, iso_date2)
    return utlpy.mysql_fecthall(connection, qry, params)


def insert_logo_cenace(ws) :    
    img = openpyxl.drawing.image.Image(LOGO_CENACE_PATH)
    ws.cell(row=2, column=2)
    ws.add_image(img)
    return

def style_cell_title(cell) :
    utlpy.pyxl_center_borderall_fill_cell(cell, fill=True, center=True, border=True, fgColor="FCE4D6")  
    return

def style_cell_header(cell):    
    utlpy.pyxl_center_borderall_fill_cell(cell, fill=True, center=True, border=True, fgColor="9BC2E6")  
    return

def style_cell_dark(cell):    
    utlpy.pyxl_center_borderall_fill_cell(cell, fill=True, center=False, border=False, fgColor="808080")  
    return

def style_cell_medicion_hr(cell):    
    utlpy.pyxl_center_borderall_fill_cell(cell, fill=False, center=True, border=True, fgColor="FFFFFF")  
    return


def fn_lookup_medicion(ds_medicion, id_equipo, date, hr, cveZona, banco):
    result = -1    
    try:
        iso_date = "{:0>4}-{:0>2}-{:0>2} {:0>2}:00:00".format(date.year, date.month, date.day, hr)    
        lst_find = [tup for tup in ds_medicion if tup[0] == iso_date]
        if(len(lst_find) >= 1) :
            result = lst_find[0][1] or -1
    except Exception as e:
        result = -1
        utlpy.println('fn_lookup_medicion.error>>> ' + str(e))        
    return result

def fn_add_date_and_hr_rows (start_date, final_date, ws) :
    row = 0
    day = 0    
    while (day < 366) :        
        date = start_date + datetime.timedelta(days=day)            
        if(date > final_date) :
            break                
        for hr in range(1, 25) :
            row = row + 1
            style_cell_medicion_hr(ws.cell(row = ZERO_ROW + row, column = 1, value = date.year))
            style_cell_medicion_hr(ws.cell(row = ZERO_ROW + row, column = 2, value = date.month))
            style_cell_medicion_hr(ws.cell(row = ZERO_ROW + row, column = 3, value = date.day))
            style_cell_medicion_hr(ws.cell(row = ZERO_ROW + row, column = 4, value = hr))                
            style_cell_dark(ws.cell(row = ZERO_ROW + row, column = 5, value = ""))
        #end for            
        day = day + 1
    #end while (day)
    return 


def fn_add_medicion_hr (connection, ws, str_anio, rs_banco, idx_banco, cveZona, banco):    
    row = 0
    day = 0
    anio = int(str_anio)    
    start_date = datetime.datetime(anio, 1, 1)
    final_date = datetime.datetime(anio, 12, 31)
    
    id_equipo = rs_banco["id_catalogo_equipos"]
    ds_medicion = fn_get_ds_medhr_by_date_range(connection, id_equipo, start_date, final_date)
    
    if(len(ds_medicion) <= 0) :
        return
    
    last_tuple = ds_medicion[-1]
    last_tuple_d = last_tuple[2]    
    last_tuple_dt = datetime.datetime(last_tuple_d.year, last_tuple_d.month, last_tuple_d.day)  
    
    if(idx_banco == 1) :
        fn_add_date_and_hr_rows(start_date, final_date, ws)
        
    while (day < 366) :        
        date = start_date + datetime.timedelta(days=day)                
        if(date > last_tuple_dt or date > final_date) :
            break            
        for hr in range(1, 25) :
            row = row + 1                        
            medicion = fn_lookup_medicion(ds_medicion, id_equipo, date, hr, cveZona, banco)            
            if(medicion != -1) :
                style_cell_medicion_hr(ws.cell(row = ZERO_ROW + row, column = ZERO_COL_BANCO + idx_banco, value = medicion))
        #end for        
        day = day + 1
    #end while (day)
        
    return

def fn_create_rpt_by_zona(connection, anio, cveDivision, wb, rs_zona) :       
    
    cveZona = rs_zona["claveZona"]        
    ws = wb.copy_worksheet(wb["ZONA"])
    ws.title = "{}{}".format(cveDivision,cveZona)
    
    insert_logo_cenace(ws)
    ws.cell(row = 3, column = 6, value = rs_zona["Abreviatura"])
    ws.cell(row = 4, column = 6, value = rs_zona["Nombre"])
    
    idx_banco = 0
    ds_banco_sub = fn_get_ds_bancos(connection, cveDivision, cveZona)
    for rs in ds_banco_sub :        
        idx_banco = idx_banco + 1
        utlpy.println("{}-{} - Banco:{}".format(cveDivision, cveZona, rs["Nombre"]))
        
        if(idx_banco > 1) :
            style_cell_title(ws.cell(row = 6, column = ZERO_COL_BANCO + idx_banco, value = ""))
        
        style_cell_header(ws.cell(row = 7, column = ZERO_COL_BANCO + idx_banco, value = rs["Nombre"]))
        style_cell_header(ws.cell(row = 8, column = ZERO_COL_BANCO + idx_banco, value = rs["nodeserie"]))
        style_cell_header(ws.cell(row = 9, column = ZERO_COL_BANCO + idx_banco, value = rs["relacionTension"]))
        style_cell_header(ws.cell(row = 10, column = ZERO_COL_BANCO + idx_banco, value = utlpy.to_number(rs["numerobanco"])))
        style_cell_header(ws.cell(row = 11, column = ZERO_COL_BANCO + idx_banco, value = utlpy.to_number(rs["capacidad3mva"])))
        style_cell_header(ws.cell(row = 12, column = ZERO_COL_BANCO + idx_banco, value = rs["numCirc"]))            
        
        fn_add_medicion_hr(connection, ws, anio, rs, idx_banco, cveZona, rs["Nombre"])
    #end-for
    
    return

def proc_genera_xls_by_zona (anio, cveDivision, cveZona, rs_zona, REPOXLS_PATH) :
    
    start_time = time.time()    
    prid = utlpy.btc_gen_prid(PR_NAME, "{}{}{}".format(cveDivision, cveZona, anio))
    
    dest_fname = "{}\{}_{}{}{}.xlsx".format(REPOXLS_PATH, PR_NAME, anio, cveDivision, cveZona)
    
    connection = None
    try:
        connection = pymysql.connect(host=CN["host"], user=CN["user"], passwd=CN["passwd"], database=CN["database"]) 
        utlpy.btc_insert(connection, prid, PR_NAME, "INICIADO", "", cveDivision, cveZona, "", "", anio, "*")
        
        utlpy.println("proc_genera_xls_by_zona:{} thread started & running (...)".format(cveZona))
        
        shutil.copy(TEMPLATE_PATH, dest_fname)
        wb = load_workbook(dest_fname)
        
        fn_create_rpt_by_zona(connection, anio, cveDivision, wb, rs_zona)
        
        wb.remove(wb["ZONA"])
        wb.save(dest_fname) 
        
        elapsed_time = time.time() - start_time
        elapsed_time_fmt = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))
        
        message = "DURACION {}".format(str(elapsed_time_fmt))
        utlpy.println("proc_genera_xls_by_zona:{} thread completed. Total time: {}".format(cveZona, elapsed_time_fmt))
        utlpy.btc_insert(connection, prid, PR_NAME, "COMPLETADO", message, cveDivision, cveZona, "", "", anio, "*")
        
    except Exception as e:
        message = str(e)
        utlpy.println('proc_genera_xls_by_zona.error>>> ' + message)        
        utlpy.btc_insert(connection, prid, PR_NAME, "ERROR", message, cveDivision, cveZona, "", "", anio, "*")
        
    finally:
        if connection:
            connection.close()
            
    return

def proc_genera_xls_by_div(cveDivision, anio, REPOXLS_PATH) :    
    
    utlpy.println("proc_genera_xls_by_div: >>> main thread started >>>")
    
    connection = None
    try:
        connection = pymysql.connect(host=CN["host"], user=CN["user"], passwd=CN["passwd"], database=CN["database"]) 
        lst_threads = []
        ds_zonas = utlpy.fn_get_ds_zonas(connection, cveDivision)
        connection.close()
        connection = None
        
        for rs_zona in ds_zonas :
            cveZona = rs_zona["claveZona"]
            t = Thread(target = proc_genera_xls_by_zona, args = (anio, cveDivision, cveZona, rs_zona, REPOXLS_PATH))
            lst_threads.append(t)
        
        [t.start() for t in lst_threads]
        [t.join() for t in lst_threads]
            
        zip_file_name = "{}_{}{}".format(PR_NAME, anio, cveDivision)
        utlpy.create_parent_zip_from_dir(REPOXLS_PATH, zip_file_name)
        
    except Exception as e:
        utlpy.println('proc_genera_xls_by_div.error>>> {}'.format(str(e)))
    finally:
        if connection:
            connection.close()
        
    print("proc_genera_xls_by_div: >>> main thread completed >>>")
    return


##########################
# main
##########################

if __name__ == '__main__' :
    
    try:    
        print("[__main__ thread started]")
        anio = "2018"
        lst_divisiones = ['DA','DB','DC','DD','DF','DG','DJ','DK','DL','DM','DN','DP','DU','DV','DW','DX']
        
        lst_threads = []
        for division in lst_divisiones :
            REPOXLS_PATH = "{}\{}\{}\{}".format(REPOXLS_BASE_PATH, PR_NAME, anio,  division)
            utlpy.create_dir_if_not_exists(REPOXLS_PATH)
            t = Thread(target = proc_genera_xls_by_div, args = (division, anio, REPOXLS_PATH))
            lst_threads.append(t)        
            
        [t.start() for t in lst_threads]
        [t.join() for t in lst_threads]
            
    except Exception as e:
        utlpy.println('__main__ thread.error>>> ' + str(e))
        
    print("[__main__ thread completed]")

        
    